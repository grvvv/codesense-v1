# projects/views.py
from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import HttpResponse
from rest_framework import status
from local.auth_app.permissions.decorators import require_permission
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from ..models.finding_models import FindingModel
 
 
class FindingListCreateView(APIView):
    @require_permission("view_findings")
    def get(self, request, scan_id):
        try:
            page = int(request.query_params.get("page", 1))
            limit = int(request.query_params.get("limit", 10))
 
            findings = FindingModel.find_by_scan(scan_id=scan_id, page=page, limit=limit)
            if not findings:
                return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
            return Response(findings, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
 
 
class ExportFindingView(APIView):
    @require_permission("create_report")
    def get(self, request, scan_id):
        try:
            findings = FindingModel.find_all_by_scan(scan_id=scan_id)
            if not findings:
                return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
 
            # Create workbook & sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Findings"
 
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border_style = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin")
            )
            code_font = Font(name="Consolas")
            code_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
 
            # Headers
            headers = [
                "Sr No", "Title", "CWE", "Severity", "CVSS Score",
                "Code Snip", "Description", "File", "Reference", "Created At"
            ]
            ws.append(headers)
 
            # Style headers
            for col_num, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border_style
 
            # Data rows
            for index, f in enumerate(findings, start=1):
                created_at_raw = f.get("created_at")
                created_at_str = ""

                if isinstance(created_at_raw, datetime):
                    created_at_str = created_at_raw.strftime("%Y-%m-%d %H:%M:%S")
                elif isinstance(created_at_raw, str):
                    try:
                        # Parse ISO string back to datetime
                        dt = datetime.fromisoformat(created_at_raw.replace("Z", "+00:00"))
                        created_at_str = dt.strftime("%Y-%m-%d %H:%M:%S")
                    except Exception:
                        created_at_str = created_at_raw

                row_data = [
                    index,
                    f.get("title", ""),
                    f.get("cwe", ""),
                    f.get("severity", ""),
                    f.get("cvss_score", ""),
                    f.get("code_snip", ""),
                    f.get("description", ""),
                    f.get("file_path", ""),
                    f.get("reference", ""),
                    created_at_str
                ]
                ws.append(row_data)
 
                # Style each cell
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=index + 1, column=col_num)
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.border = border_style
 
                    # Severity column coloring
                    if col_num == 4:
                        severity = str(value).lower()
                        if severity == "critical":
                            cell.fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
                        elif severity == "high":
                            cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
                        elif severity == "medium":
                            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                        elif severity == "low":
                            cell.fill = PatternFill(start_color="A8D08D", end_color="A8D08D", fill_type="solid")
                        elif severity == "info":
                            cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
 
                    # Code snippet styling
                    if col_num == 6:
                        cell.font = code_font
                        cell.fill = code_fill
 
            # Auto column width
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
 
            # Response
            response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            response["Content-Disposition"] = f'attachment; filename="scan_{scan_id}_findings.xlsx"'
            wb.save(response)
 
            return response
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
 
 
class FindingDetailView(APIView):
    @require_permission("view_findings")
    def get(self, request, finding_id):
        finding = FindingModel.find_by_id(finding_id=finding_id)
        if not finding:
            return Response({"error": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        return Response(finding, status=status.HTTP_200_OK)
 
    @require_permission("validate_finding")
    def patch(self, request, finding_id):
        updated_finding = FindingModel.toggle_approved(finding_id=finding_id)
        if not updated_finding:
            return Response({"error": "Not found or not updated"}, status=status.HTTP_404_NOT_FOUND)
        return Response(updated_finding, status=status.HTTP_200_OK)
 
    @require_permission("delete_finding")
    def delete(self, request, finding_id):
        FindingModel.soft_delete(finding_id=finding_id)
        return Response(status=status.HTTP_204_NO_CONTENT)